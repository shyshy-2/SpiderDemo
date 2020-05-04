#-*- coding = utf-8 -*-
#@Time : XXXXXXXXXXXXXXXX
#@Author : shy-2
#@File : word_cloud.py
#@Software : PyCharm

import openpyxl
from wordcloud import WordCloud

#读取数据
wb = openpyxl.load_workbook('data.xlsx')

#获取工作表
ws = wb['国内疫情']
frequency_in = {}
for row in ws.values:
    if row[0] == '省份':
        pass
    else:
        frequency_in[row[0]] = float(row[1])

# print(frequency)

wordcloud = WordCloud(font_path='C:/Windows/SIMLI.TTF',background_color="white",
                      width=1920,height=1080)

#根据确诊病例的数目生成词云
wordcloud.generate_from_frequencies(frequency_in)

#保存词云
wordcloud.to_file("wordcloud.png")

frequency_out = {}

# print(wb.sheetnames)
sheet_names = wb.sheetnames
for each in sheet_names:
    if "洲" in each:
        ws = wb[each]
        for row in ws.values:
            # print(row)
            if row[0] == "国家":
                pass
            else:
                frequency_out[row[0]] = float(row[1])

#根据确诊病例的数目生成词云
wordcloud.generate_from_frequencies(frequency_out)
#保存词云
wordcloud.to_file("wordcloud.png")